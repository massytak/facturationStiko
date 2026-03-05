"""
STIKO TRANS — Backend FastAPI
=============================
Endpoints :
  POST /api/parse-pdf     → extrait les lignes du relevé PDF
  POST /api/generate-excel → génère le fichier Excel facture
"""

import re
import io
from datetime import datetime, timedelta
from typing import Optional

import pdfplumber
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = FastAPI(title="STIKO TRANS API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:5173"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ──────────────────────────────────────────────
# MODÈLES PYDANTIC
# ──────────────────────────────────────────────

class LigneTransport(BaseModel):
    date_str: str
    commande: str
    container: str
    quantite: int
    immob: float
    total_ht: float
    tva: float
    total_ttc: float
    is_exo: bool

class ParseResult(BaseModel):
    lignes: list[LigneTransport]
    date_debut: str
    date_fin: str
    facture_num_detecte: Optional[str]
    total_ht: float
    total_tva: float
    total_ttc: float

class GenerateRequest(BaseModel):
    lignes: list[LigneTransport]
    facture_num: str
    facture_date: str          # ISO format YYYY-MM-DD
    client_nom: str
    client_adresse: str = "BOULEVARD JULES DURAND"
    client_cp_ville: str = "76600 LE HAVRE"
    client_email: str = "admin1@transfret-me.fr"
    client_siret: str = "34401503700072"

# ──────────────────────────────────────────────
# PARSING
# ──────────────────────────────────────────────

def parse_pdf_text(text: str) -> list[dict]:
    """
    Parse le texte complet du relevé TRANSFRET.

    Structure ligne principale :
      DD/MM/YY  H XX XX XXXX  CONTAINER  ...DESIG...  QTE,00  PU,00  MONTANT,xx  CODE_TVA
    Ligne immobilisation (optionnelle) :
      FRAIS D IMMOBILISATION  NB,00  40,00  TOTAL,00
    Ligne total HT :
      IMPORT IM4-xxx  HT,xx  TVA,xx  TTC,xx
      EXPORT EXO-xxx  HT,xx  TTC,xx
    """

    # Ligne principale transport
    line_re = re.compile(
        r"(\d{2}/\d{2}/\d{2,4})"
        r"\s+(H\s*\d{2}\s*\d{2}\s*\d{4})"
        r"\s+([A-Z]{4}\d{5,8})"
        r"\s+.+?"
        r"\s+(\d{1,5}),00"     # QTE (jusqu'à 5 chiffres)
        r"\s+\d{1,4},00"       # PU (ignoré)
        r"\s+[\d]+,\d{2}"      # montant ligne (ignoré)
        r"\s+(\d)\b"           # code TVA : 0=exo 4=20%
    )

    # Total HT sur ligne IMPORT/EXPORT
    total_re = re.compile(
        r"(?:IMPORT\s+\S+-\S*|EXPORT\s+EXO-.*?)\s+(\d+,\d{2})"
    )

    # Immobilisation
    immo_re = re.compile(
        r"FRAIS D IMMOBILISATION\s+[\d,]+\s+40,00\s+(\d+),00"
    )

    all_totaux = [
        {"pos": m.start(), "ht": float(m.group(1).replace(",", "."))}
        for m in total_re.finditer(text)
    ]
    all_immos = [
        {"pos": m.start(), "total": float(m.group(1))}
        for m in immo_re.finditer(text)
    ]

    matches = []
    for m in line_re.finditer(text):
        matches.append({
            "start":    m.start(),
            "date":     m.group(1),
            "cmd":      re.sub(r"\s+", " ", m.group(2)).strip(),
            "ctr":      m.group(3),
            "qte":      int(m.group(4)),
            "tva_code": m.group(5),
        })

    results = []
    for i, match in enumerate(matches):
        seg_start = match["start"]
        seg_end   = matches[i + 1]["start"] if i + 1 < len(matches) else len(text)
        is_exo    = match["tva_code"] == "0"

        block_tot  = [x for x in all_totaux if seg_start <= x["pos"] < seg_end]
        total_ht   = block_tot[0]["ht"] if block_tot else 0.0

        block_immo = [x for x in all_immos if seg_start <= x["pos"] < seg_end]
        immob      = sum(x["total"] for x in block_immo)

        dp  = match["date"].split("/")
        yr  = "20" + dp[2] if len(dp[2]) == 2 else dp[2]

        tva      = 0.0 if is_exo else round(total_ht * 0.2, 2)
        total_ttc = round(total_ht + tva, 2)

        results.append({
            "date_str":  f"{dp[0]}/{dp[1]}/{yr}",
            "commande":  match["cmd"],
            "container": match["ctr"],
            "quantite":  match["qte"],
            "immob":     immob,
            "total_ht":  total_ht,
            "tva":       tva,
            "total_ttc": total_ttc,
            "is_exo":    is_exo,
        })

    return results


# ──────────────────────────────────────────────
# GÉNÉRATION EXCEL
# ──────────────────────────────────────────────

def build_excel(req: GenerateRequest) -> bytes:
    lignes       = req.lignes
    facture_num  = req.facture_num
    facture_date = datetime.fromisoformat(req.facture_date)
    client_nom   = req.client_nom

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "transfret-"

    BLUE    = "FF4FB8F9"
    RED     = "FFFF0000"
    WHITE   = "FFFFFFFF"
    GREY_BG = "FFF2F5F7"
    DARK_BG = "FF2C3E50"

    def fill(rgb):   return PatternFill("solid", fgColor=rgb)
    def fnt(bold=False, size=10, color=None, italic=False):
        kw = dict(name="Arial", bold=bold, size=size, italic=italic)
        if color: kw["color"] = color
        return Font(**kw)
    def aln(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def bdr():
        s = Side(border_style="thin", color="FFD0D5DD")
        return Border(bottom=s)

    for col, w in {"A":2,"B":28,"C":16,"D":10,"E":14,"F":11,"G":16,"H":13,"I":18}.items():
        ws.column_dimensions[col].width = w
    for r, h in {1:8,2:34,3:4,4:8,5:16,6:16,7:16,8:16,9:16,10:22,
                 11:8,12:16,13:28,14:16,15:28,16:16,17:22,18:6}.items():
        ws.row_dimensions[r].height = h
    for r in range(19, 19 + len(lignes) + 15):
        ws.row_dimensions[r].height = 24

    # Titre
    ws["B2"] = "STIKO TRANS"; ws["B2"].font = fnt(bold=True, size=22, color=BLUE); ws["B2"].alignment = aln()
    ws.merge_cells("G2:I2"); ws["G2"] = "Facture"; ws["G2"].font = fnt(bold=True, size=16); ws["G2"].alignment = aln("right")
    for col in "BCDEFGHI": ws[col+"3"].fill = fill(BLUE)

    # Émetteur
    for row, txt in [(5,"SAS STIKO TRANS"),(6,"Société au capital de 14.000 euros"),
                     (7,"10 RUE DE PENTHIEVRE 75008 PARIS"),
                     (8,"Numéro siret : 98097798700018"),(9,"Numéro tva : FR48980977987")]:
        ws["B"+str(row)] = txt; ws["B"+str(row)].font = fnt(bold=True, size=10, color=RED)

    # Méta
    week_num = facture_date.isocalendar()[1]
    for addr, label, val in [("G5","Numéro de facture :",facture_num),
                              ("G7","Échéance :","15 JOURS"),("I8",None,f"S{week_num}")]:
        if label: ws[addr] = label; ws[addr].font = fnt(size=10)
        ws["I"+addr[1:]] = val if not label else val
        ws["I"+addr[1:]].font = fnt(bold=True, size=10); ws["I"+addr[1:]].fill = fill(GREY_BG); ws["I"+addr[1:]].alignment = aln("right")
    ws["G6"] = "Date émission :"; ws["G6"].font = fnt(size=10)
    ws["I6"] = facture_date; ws["I6"].number_format = "DD/MM/YYYY"; ws["I6"].font = fnt(size=10); ws["I6"].fill = fill(GREY_BG); ws["I6"].alignment = aln("right")
    ws["I8"] = f"S{week_num}"; ws["I8"].font = fnt(bold=True, size=10); ws["I8"].fill = fill(GREY_BG); ws["I8"].alignment = aln("center")

    # Sections client
    ws["B10"] = "Facturer à :"; ws["B10"].font = fnt(bold=True,size=10,color=WHITE); ws["B10"].fill = fill(DARK_BG)
    ws.merge_cells("G10:I10"); ws["G10"] = "Facturé au client :"; ws["G10"].font = fnt(bold=True,size=10,color=WHITE); ws["G10"].fill = fill(DARK_BG)
    d1 = lignes[0].date_str if lignes else ""; d2 = lignes[-1].date_str if lignes else ""
    ws["B13"] = "Date des prestations effectuées :"; ws["B14"] = f"Du {d1} au {d2}"
    for row, label, value in [(12,"Nom",client_nom),(13,"Adresse",req.client_adresse),
                               (14,"CP/Ville",req.client_cp_ville),(15,"Email",req.client_email),
                               (16,"SIRET",req.client_siret)]:
        ws["G"+str(row)] = label; ws["G"+str(row)].font = fnt(size=10)
        ws["I"+str(row)] = value; ws["I"+str(row)].font = fnt(size=10); ws["I"+str(row)].fill = fill(GREY_BG)

    # En-tête tableau
    for addr, label in [("B17","Date"),("C17","Commande"),("D17","Quantité"),
                        ("E17","HEURE D'ATTENTE"),("F17","MULTISTOP"),
                        ("G17","Prix unitaire HT"),("H17","PRIX TVA"),("I17","Prix total HT")]:
        ws[addr] = label; ws[addr].font = fnt(bold=True,size=10,color=WHITE)
        ws[addr].fill = fill(DARK_BG); ws[addr].alignment = aln("center")

    # Données
    first_row = 19
    EUR = '#,##0.00 [$€]'
    for i, ligne in enumerate(lignes):
        r = first_row + i
        try:
            dp = ligne.date_str.split("/")
            date_obj = datetime(int(dp[2]), int(dp[1]), int(dp[0]))
        except Exception:
            date_obj = facture_date
        ws[f"B{r}"] = date_obj; ws[f"B{r}"].number_format = "DD/MM/YYYY"; ws[f"B{r}"].font = fnt(size=10); ws[f"B{r}"].fill = fill(GREY_BG); ws[f"B{r}"].border = bdr(); ws[f"B{r}"].alignment = aln("center")
        ws[f"C{r}"] = ligne.commande; ws[f"C{r}"].font = fnt(size=10); ws[f"C{r}"].fill = fill(GREY_BG); ws[f"C{r}"].border = bdr(); ws[f"C{r}"].alignment = aln("center")
        ws[f"D{r}"] = ligne.quantite; ws[f"D{r}"].font = fnt(size=10); ws[f"D{r}"].fill = fill(GREY_BG); ws[f"D{r}"].alignment = aln("center"); ws[f"D{r}"].border = bdr()
        ws[f"E{r}"] = ligne.immob; ws[f"E{r}"].number_format = EUR; ws[f"E{r}"].font = fnt(size=10); ws[f"E{r}"].fill = fill(GREY_BG); ws[f"E{r}"].border = bdr()
        ws[f"F{r}"] = 0; ws[f"F{r}"].font = fnt(size=10); ws[f"F{r}"].fill = fill(GREY_BG); ws[f"F{r}"].border = bdr()
        ws[f"G{r}"] = "forfait"; ws[f"G{r}"].font = fnt(bold=True,size=10); ws[f"G{r}"].fill = fill(GREY_BG); ws[f"G{r}"].alignment = aln("center"); ws[f"G{r}"].border = bdr()
        ws[f"H{r}"] = f"=I{r}*0.2" if not ligne.is_exo else 0
        ws[f"H{r}"].font = fnt(size=10); ws[f"H{r}"].fill = fill(GREY_BG); ws[f"H{r}"].number_format = EUR; ws[f"H{r}"].alignment = aln("right"); ws[f"H{r}"].border = bdr()
        ws[f"I{r}"] = ligne.total_ht; ws[f"I{r}"].font = fnt(size=10); ws[f"I{r}"].fill = fill(GREY_BG); ws[f"I{r}"].number_format = EUR; ws[f"I{r}"].alignment = aln("right"); ws[f"I{r}"].border = bdr()

    # Totaux
    last_row = first_row + len(lignes) - 1
    r_ht = last_row + 2; r_tva = r_ht + 1; r_somme = r_tva + 1
    r_bank = r_somme + 4; r_foot = r_bank + 5

    ws.merge_cells(f"D{r_ht}:G{r_ht}"); ws[f"D{r_ht}"] = "Prix total HT :"; ws[f"D{r_ht}"].font = fnt(bold=True,size=10); ws[f"D{r_ht}"].alignment = aln("right")
    ws[f"I{r_ht}"] = f"=SUM(I{first_row}:I{last_row})"; ws[f"I{r_ht}"].font = fnt(bold=True,size=11); ws[f"I{r_ht}"].number_format = EUR; ws[f"I{r_ht}"].alignment = aln("right")
    ws.merge_cells(f"D{r_tva}:G{r_tva}"); ws[f"D{r_tva}"] = "Prix total TVA :"; ws[f"D{r_tva}"].font = fnt(bold=True,size=10); ws[f"D{r_tva}"].alignment = aln("right")
    ws[f"I{r_tva}"] = f"=SUM(H{first_row}:H{last_row})"; ws[f"I{r_tva}"].font = fnt(bold=True,size=11); ws[f"I{r_tva}"].number_format = EUR; ws[f"I{r_tva}"].alignment = aln("right")
    ws.merge_cells(f"D{r_somme}:G{r_somme+1}"); ws[f"D{r_somme}"] = "SOMME FINALE À PAYER :"; ws[f"D{r_somme}"].font = fnt(bold=True,size=11,color=BLUE); ws[f"D{r_somme}"].alignment = aln("right")
    ws.merge_cells(f"H{r_somme}:I{r_somme+1}")
    ws[f"H{r_somme}"] = f"=I{r_ht}+I{r_tva}"
    ws[f"H{r_somme}"].font = Font(name="Arial",bold=True,size=14,color=BLUE)
    ws[f"H{r_somme}"].number_format = EUR; ws[f"H{r_somme}"].fill = fill(GREY_BG); ws[f"H{r_somme}"].alignment = aln("center")
    ws[f"H{r_somme}"].border = Border(top=Side(border_style="medium",color=BLUE),bottom=Side(border_style="medium",color=BLUE),
                                      left=Side(border_style="medium",color=BLUE),right=Side(border_style="medium",color=BLUE))
    ech = facture_date + timedelta(days=14)
    ws[f"B{r_somme}"] = f"Date d'échéance : {ech.strftime('%d/%m/%Y')}"
    ws[f"B{r_somme}"].font = fnt(bold=True, size=10)
    ws.merge_cells(f"B{r_somme}:C{r_somme}")
    ws[f"B{r_bank}"] = "COORDONNÉES BANCAIRES :"; ws[f"B{r_bank}"].font = fnt(bold=True,size=10)
    ws[f"B{r_bank+1}"] = "IBAN : FR76 2823 3000 0179 4403 5112 530"; ws[f"B{r_bank+1}"].font = fnt(size=10)
    ws[f"B{r_bank+2}"] = "BIC : REVOFRP2"; ws[f"B{r_bank+2}"].font = fnt(size=10)
    ws.merge_cells(f"B{r_foot}:I{r_foot}")
    ws[f"B{r_foot}"] = "SIREN 980977987 | NAF 49.41B | TVA FR48980977987"
    ws[f"B{r_foot}"].font = fnt(size=8,color="FF888888"); ws[f"B{r_foot}"].alignment = aln("center")

# ── Auto-ajustement de la largeur de toutes les colonnes
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
        if col_letter == "B":
            ws.column_dimensions["B"].width = 17   # ← largeur fixe pour B
        else:
            ws.column_dimensions[col_letter].width = max_len + 4
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ──────────────────────────────────────────────
# ROUTES
# ──────────────────────────────────────────────

@app.get("/")
def root():
    return {"status": "ok", "service": "STIKO TRANS API v1.0"}


@app.post("/api/parse-pdf", response_model=ParseResult)
async def parse_pdf(file: UploadFile = File(...)):
    """Parse le relevé PDF et retourne les lignes de transport extraites."""
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Fichier PDF requis")

    content = await file.read()
    buf = io.BytesIO(content)

    try:
        full_text = ""
        with pdfplumber.open(buf) as pdf:
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"
    except Exception as e:
        raise HTTPException(500, f"Erreur lecture PDF : {e}")

    lignes_raw = parse_pdf_text(full_text)

    if not lignes_raw:
        raise HTTPException(422, "Aucune ligne de transport détectée dans ce relevé")

    lignes = [LigneTransport(**l) for l in lignes_raw]

    # Numéro de facture auto-détecté
    num_match = re.search(r"\bST\s*(\d{3,4})\b", full_text, re.IGNORECASE)
    facture_num = "ST" + num_match.group(1) if num_match else None

    total_ht  = round(sum(l.total_ht  for l in lignes), 2)
    total_tva = round(sum(l.tva       for l in lignes), 2)
    total_ttc = round(sum(l.total_ttc for l in lignes), 2)

    return ParseResult(
        lignes=lignes,
        date_debut=lignes[0].date_str,
        date_fin=lignes[-1].date_str,
        facture_num_detecte=facture_num,
        total_ht=total_ht,
        total_tva=total_tva,
        total_ttc=total_ttc,
    )


@app.post("/api/generate-excel")
async def generate_excel(req: GenerateRequest):
    """Génère et retourne le fichier Excel facture."""
    if not req.lignes:
        raise HTTPException(400, "Aucune ligne fournie")

    try:
        excel_bytes = build_excel(req)
    except Exception as e:
        raise HTTPException(500, f"Erreur génération Excel : {e}")

    fname = f"FACTURE_STIKO_TRANS_{req.facture_num}_{req.facture_date.replace('-','')}.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )
